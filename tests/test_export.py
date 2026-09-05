# tests/test_export.py
import csv
from datetime import datetime
from decimal import Decimal

import duckdb
import pytest
from openpyxl import load_workbook

from od_cpd import export
from od_cpd.dbio import sql_literal


def test_write_csv(tmp_path):
    con = duckdb.connect(":memory:")
    con.execute("CREATE TABLE t AS SELECT 1 AS a, 'x' AS b")
    out = export.write_csv(con, "SELECT * FROM t", tmp_path / "o.csv")
    assert out.exists()
    with out.open(newline="") as stream:
        assert list(csv.reader(stream)) == [["a", "b"], ["1", "x"]]


def test_write_csv_trailing_comment(tmp_path):
    # The text-casting wrapper must tolerate a trailing SQL comment.
    con = duckdb.connect(":memory:")
    con.execute("CREATE TABLE t AS SELECT 1 AS a, 'x' AS b")
    out = export.write_csv(con, "SELECT * FROM t -- trailing comment", tmp_path / "o.csv")
    assert out.exists()
    with out.open(newline="") as stream:
        assert list(csv.reader(stream)) == [["a", "b"], ["1", "x"]]


def test_write_csv_quoted_path_and_values(tmp_path):
    with duckdb.connect(":memory:") as con:
        values = ('comma, quote " and\nnewline — café', None, Decimal("12.34"))
        con.execute("CREATE TABLE t (text VARCHAR, missing VARCHAR, amount DECIMAL(8, 2))")
        con.execute("INSERT INTO t VALUES (?, ?, ?)", values)
        out = export.write_csv(con, "SELECT * FROM t", tmp_path / "author's export.csv")
    with out.open(encoding="utf-8", newline="") as stream:
        assert list(csv.reader(stream)) == [
            ["text", "missing", "amount"],
            [values[0], "", "12.34"],
        ]


def test_write_csv_streams_full_result(tmp_path):
    with duckdb.connect(":memory:") as con:
        out = export.write_csv(con, "SELECT n FROM range(10001) AS r(n)", tmp_path / "all.csv")
    with out.open(newline="") as stream:
        rows = list(csv.reader(stream))
    assert rows == [["n"], *[[str(n)] for n in range(10001)]]


def test_write_csv_preserves_null_and_empty_string(tmp_path):
    with duckdb.connect(":memory:") as con:
        out = export.write_csv(
            con, "SELECT NULL::VARCHAR AS missing, '' AS empty", tmp_path / "nulls.csv",
        )
        assert out.read_text() == 'missing,empty\n,""\n'
        # Treat an unquoted empty field as NULL and a quoted empty field as ''.
        assert con.execute(
            "SELECT * FROM read_csv(?, all_varchar=true, allow_quoted_nulls=false)",
            [str(out)],
        ).fetchall() == [(None, "")]


@pytest.mark.parametrize("query", [
    """SELECT INTERVAL '1 month 2 days 03:04:05.123456' AS elapsed,
              true AS flag, false AS other_flag, [1, NULL, 3] AS numbers,
              {'key': 'value', 'n': NULL} AS details,
              123.450::DECIMAL(10, 3) AS amount, DATE 'infinity' AS endpoint,
              TIMESTAMP_NS '2026-01-01 00:00:00.123456789' AS instant""",
    'SELECT 1 AS "same,label", 2 AS "same,label", 3 AS "quote""header" WHERE false',
])
def test_write_csv_matches_duckdb_text_and_headers(tmp_path, query):
    baseline = tmp_path / "baseline.csv"
    with duckdb.connect(":memory:") as con:
        # Trusted synthetic SQL and a temporary destination only: COPY is the
        # reference format here, never part of the production export path.
        con.execute(f"COPY ({query}) TO {sql_literal(str(baseline))} (HEADER, DELIMITER ',')")
        out = export.write_csv(con, query, tmp_path / "application.csv")
    with baseline.open(newline="") as reference, out.open(newline="") as actual:
        assert list(csv.reader(actual)) == list(csv.reader(reference))


@pytest.mark.parametrize("value", ["=1+1", "+1+1", "-1+1", "@SUM(A1)", "#N/A"])
def test_csv_preserves_formula_like_text_without_modification(tmp_path, value):
    with duckdb.connect(":memory:") as con:
        out = export.write_csv(con, f'SELECT {sql_literal(value)} AS "=1+1"', tmp_path / "text.csv")
    with out.open(newline="") as stream:
        assert list(csv.reader(stream)) == [["=1+1"], [value]]


def test_xlsx_strings_headers_and_methodology_are_literal(tmp_path):
    query = '''SELECT '=1+1' AS "=2+2", '#N/A' AS error_text,
                      '+1+1' AS plus_text, '-1+1' AS minus_text, '@SUM(A1)' AS at_text'''
    provenance = {"=3+3": "=4+4", "#N/A": "#VALUE!"}
    with duckdb.connect(":memory:") as con:
        out = export.write_xlsx(con, query, provenance, tmp_path / "literal.xlsx")
    wb = load_workbook(out, data_only=False)
    try:
        assert list(wb["data"].values) == [
            ("=2+2", "error_text", "plus_text", "minus_text", "at_text"),
            ("=1+1", "#N/A", "+1+1", "-1+1", "@SUM(A1)"),
        ]
        assert list(wb["methodology"].values) == list(provenance.items())
        for sheet in wb:
            assert all(cell.data_type == "s" for row in sheet for cell in row)
    finally:
        wb.close()


def test_xlsx_preserves_numbers_dates_and_nested_values(tmp_path):
    query = """SELECT 123.45::DECIMAL(8,2) AS amount, DATE '2026-01-02' AS day,
                      true AS flag, ['=1+1', 'x'] AS items, {'formula': '=2+2'} AS detail"""
    with duckdb.connect(":memory:") as con:
        out = export.write_xlsx(con, query, {}, tmp_path / "typed.xlsx")
    wb = load_workbook(out, data_only=False)
    try:
        row = list(wb["data"].rows)[1]
        assert [cell.value for cell in row] == [
            123.45, datetime(2026, 1, 2), True, "['=1+1', 'x']", "{'formula': '=2+2'}",
        ]
        assert [cell.data_type for cell in row] == ["n", "d", "b", "s", "s"]
    finally:
        wb.close()


def test_write_xlsx_trailing_comment(tmp_path):
    # write_xlsx executes the query directly (no paren wrap), so a trailing comment is
    # already harmless — this guards that it stays that way.
    con = duckdb.connect(":memory:")
    con.execute("CREATE TABLE t AS SELECT 1 AS a")
    prov = {"definition": "d", "reproduce_sql": "SELECT * FROM t"}
    out = export.write_xlsx(con, "SELECT * FROM t -- trailing comment", prov, tmp_path / "o.xlsx")
    assert out.exists()


def test_write_csv_timeout(tmp_path):
    # A pathologically slow, uncapped query must be interrupted by the timeout rather than
    # hang the server. range() cross join = ~10^18 rows; a 1s timeout fires long first.
    con = duckdb.connect(":memory:")
    slow = "SELECT count(*) FROM range(1000000000) a, range(1000000000) b"
    with pytest.raises(duckdb.Error):
        export.write_csv(con, slow, tmp_path / "o.csv", timeout=1)


def test_write_xlsx_two_sheets(tmp_path):
    con = duckdb.connect(":memory:")
    con.execute("CREATE TABLE t AS SELECT 1 AS a")
    prov = {"definition": "d", "reproduce_sql": "SELECT * FROM t"}
    out = export.write_xlsx(con, "SELECT * FROM t", prov, tmp_path / "o.xlsx")
    wb = load_workbook(out)
    try:
        assert set(wb.sheetnames) == {"data", "methodology"}
    finally:
        wb.close()
