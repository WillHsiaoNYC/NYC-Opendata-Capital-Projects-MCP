"""Regression tests for the second /code-review fix round (max effort, 2026-06-09)."""
import duckdb
import pytest

from od_cpd import export, ingest, materialize, schema
from od_cpd.tools._common import escape_like
from od_cpd.tools.budget import budget_change_from
from od_cpd.tools.inspect import get_project_budget_from
from od_cpd.tools.lifecycle import project_duration_stats_from
from od_cpd.tools.lookup import describe_field_from
from od_cpd.tools.ranking import rank_projects_from
from od_cpd.tools.resolve import resolve_from
from od_cpd.tools.schedule import schedule_breakdown_from, schedule_changes_from
from od_cpd.tools.sql import run_sql_on, validate_select
from tests.test_agency_scope import _scope_db
from tests.test_materialize_normalized import _raw


def _db():
    con = duckdb.connect(":memory:"); _raw(con); materialize.materialize_all(con)
    return con


# ---- rank_projects: every advertised filter must actually filter ------------

def test_rank_budget_applies_min_total_budget():
    con = _scope_db(duckdb.connect(":memory:"))
    r = rank_projects_from(con, entity="budget", rank_by="total_budget",
                           min_total_budget=1000)
    assert {row["fms_id"] for row in r["rows"]} == {"J2"}   # only J2 ($9000) clears it


def test_rank_budget_applies_max_total_budget():
    con = _scope_db(duckdb.connect(":memory:"))
    r = rank_projects_from(con, entity="budget", rank_by="total_budget",
                           max_total_budget=60)
    assert {row["fms_id"] for row in r["rows"]} == {"B"}    # only B ($50)


def test_rank_schedule_applies_delayed_only():
    con = duckdb.connect(":memory:"); _raw(con)
    # add an EARLY project (negative variance) that delayed_only must exclude
    con.execute(
        "INSERT INTO raw_project_detail (reporting_period, managing_agency, sponsor_agency,"
        " pid, fms_id, total_budget, current_phase) "
        "VALUES ('202601','DDC','DDC','103','G','10','Construction')")
    con.execute(
        "INSERT INTO raw_schedule_history (reporting_period, managing_agency, pid,"
        " current_phase, variance_day) VALUES ('202601','DDC','103','Construction','-30')")
    materialize.materialize_all(con)
    r = rank_projects_from(con, entity="schedule", rank_by="period_variance_days",
                           direction="bottom", delayed_only=True)
    pids = {row["pid"] for row in r["rows"]}
    assert pids == {"101"}                      # 103 (-30, earlier) filtered out
    without = rank_projects_from(con, entity="schedule", rank_by="period_variance_days",
                                 direction="bottom", delayed_only=False)
    assert "103" in {row["pid"] for row in without["rows"]}


# ---- schedule_breakdown: statistic is validated and honestly labeled --------

def test_breakdown_rejects_unknown_statistic():
    con = _db()
    r = schedule_breakdown_from(con, group_by="managing_agency",
                                metric="schedule_variance", statistic="stddev")
    assert "error" in r


def test_breakdown_default_count_is_a_real_count():
    con = _db()
    r = schedule_breakdown_from(con, group_by="managing_agency",
                                metric="schedule_variance")  # statistic defaults to count
    ddc = next(g for g in r["groups"] if g["managing_agency"] == "DDC")
    assert ddc["value"] == 1                    # one PID with a variance value, not avg=45
    assert "direction" not in ddc               # counts are unsigned magnitudes


def test_breakdown_min_statistic_supported():
    con = _db()
    r = schedule_breakdown_from(con, group_by="managing_agency",
                                metric="schedule_variance", statistic="min")
    ddc = next(g for g in r["groups"] if g["managing_agency"] == "DDC")
    assert ddc["value"] == 45 and ddc["direction"] == "later"


# ---- schedule_changes: periods validated; 'delayed' is truly between-periods

def test_schedule_changes_rejects_off_cadence_period():
    con = _db()
    r = schedule_changes_from(con, change_type="completed",
                              from_period="202512", to_period="202601")
    assert "error" in r and "01/05/09" in r["error"]


def test_schedule_changes_rejects_missing_to_period():
    con = _db()
    r = schedule_changes_from(con, change_type="completed",
                              from_period="202505", to_period="202509")
    assert "error" in r                          # 202509 not in schedule_history


def test_schedule_changes_rejects_inverted_periods():
    con = _db()
    r = schedule_changes_from(con, change_type="delayed",
                              from_period="202601", to_period="202509")
    assert "error" in r


def test_schedule_changes_delayed_excludes_already_delayed():
    con = duckdb.connect(":memory:"); schema.apply_schema(con)
    con.executemany(
        "INSERT INTO raw_project_detail (reporting_period, managing_agency, sponsor_agency,"
        " pid, fms_id, total_budget, current_phase) VALUES (?,?,?,?,?,?,?)",
        [["202509", "DDC", "DDC", "104", "H", "10", "Construction"],
         ["202601", "DDC", "DDC", "104", "H", "10", "Construction"],
         ["202509", "DDC", "DDC", "105", "I", "10", "Construction"],
         ["202601", "DDC", "DDC", "105", "I", "10", "Construction"]])
    con.executemany(
        "INSERT INTO raw_schedule_history (reporting_period, managing_agency, pid,"
        " current_phase, variance_day) VALUES (?,?,?,?,?)",
        [["202509", "DDC", "104", "Construction", "30"],   # already delayed at from
         ["202601", "DDC", "104", "Construction", "40"],
         ["202509", "DDC", "105", "Construction", "0"],    # flat at from
         ["202601", "DDC", "105", "Construction", "25"]])  # newly delayed
    materialize.materialize_all(con)
    r = schedule_changes_from(con, change_type="delayed",
                              from_period="202509", to_period="202601")
    assert {c["pid"] for c in r["changes"]} == {"105"}
    assert "note" not in r                       # from_period exists, no caveat needed


def test_schedule_changes_notes_predata_from_period():
    con = _db()
    r = schedule_changes_from(con, change_type="delayed",
                              from_period="202509", to_period="202601")
    assert "note" in r                           # 202509 predates the fixture's data


# ---- project_duration_stats: group_by works ---------------------------------

def _duration_db():
    con = duckdb.connect(":memory:"); schema.apply_schema(con)
    con.executemany(
        "INSERT INTO raw_project_detail (reporting_period, managing_agency, pid, fms_id,"
        " total_budget, current_phase, actual_design_start, actual_construction_end)"
        " VALUES (?,?,?,?,?,?,?,?)",
        [["202601", "DDC", "201", "Z", "10", "Close-out", "2020-01-01", "2023-01-01"],
         ["202601", "DPR", "301", "Y", "10", "Close-out", "2021-01-01", "2022-01-01"]])
    materialize.materialize_all(con)
    return con


def test_duration_stats_group_by_managing_agency():
    r = project_duration_stats_from(_duration_db(), group_by="managing_agency")
    assert r["group_by"] == "managing_agency"
    by = {g["managing_agency"]: g for g in r["groups"]}
    assert by["DDC"]["n_projects"] == 1 and by["DDC"]["mean_days"] > 1000
    assert by["DPR"]["n_projects"] == 1 and by["DPR"]["mean_days"] == 365
    assert r["n_projects"] == 2


def test_duration_stats_rejects_unknown_group_by():
    assert "error" in project_duration_stats_from(_duration_db(), group_by="sponsor_agency")


# ---- get_project_budget: current links only, case-insensitive id ------------

def _stale_link_db():
    con = duckdb.connect(":memory:"); schema.apply_schema(con)
    con.executemany(
        "INSERT INTO raw_project_detail (reporting_period, managing_agency, sponsor_agency,"
        " pid, fms_id, total_budget, current_phase) VALUES (?,?,?,?,?,?,?)",
        [["202509", "DDC", "DDC", "101", "A", "100", "Construction"],
         ["202509", "DDC", "DDC", "999", "A", "100", "Construction"],  # stale link
         ["202601", "DDC", "DDC", "101", "A", "100", "Construction"]])
    con.execute(
        "INSERT INTO raw_budget_history (managing_agency, fms_id, year_month_reported,"
        " total_budget, spend_to_date, budget_variance) "
        "VALUES ('DDC','A','202601','100','10','0')")
    materialize.materialize_all(con)
    return con


def test_budget_links_use_latest_link_period_only():
    r = get_project_budget_from(_stale_link_db(), "A")
    assert {s["pid"] for s in r["linked_schedules"]} == {"101"}   # 999 dropped periods ago
    assert "1:1" in r["caveat"]                                   # not a fabricated fan-out


def test_budget_lookup_is_case_insensitive():
    r = get_project_budget_from(_stale_link_db(), "a")
    assert "error" not in r
    assert r["answer"][0]["fms_id"] == "A"


def test_budget_change_fms_target_case_insensitive():
    con = _db()
    r = budget_change_from(con, target="fms:a", from_period="202509", to_period="202601")
    assert r["from_value"] == 90.0 and r["to_value"] == 100.0


# ---- budget_change: unknown metric errors instead of mislabeling ------------

def test_budget_change_rejects_unknown_metric():
    con = _db()
    r = budget_change_from(con, target="fms:A", from_period="202509",
                           to_period="202601", metric="spend_to_date")
    assert "error" in r


# ---- exports: unique paths, non-scalar cells don't crash xlsx ---------------

def test_exports_get_unique_paths(tmp_path, monkeypatch):
    monkeypatch.setenv("OD_CPD_EXPORT_DIR", str(tmp_path))
    con = duckdb.connect(":memory:")
    con.execute("CREATE TABLE t AS SELECT 1 AS a")
    p1 = export.write_csv(con, "SELECT * FROM t")
    p2 = export.write_csv(con, "SELECT * FROM t")
    assert p1 != p2 and p1.exists() and p2.exists()


def test_xlsx_handles_list_and_struct_columns(tmp_path):
    con = duckdb.connect(":memory:")
    out = export.write_xlsx(
        con, "SELECT [1, 2] AS l, {'a': 1} AS s, 'x' AS v",
        {"definition": "d"}, tmp_path / "o.xlsx")
    from openpyxl import load_workbook
    wb = load_workbook(out)
    ws = wb["data"]
    assert ws.cell(row=2, column=3).value == "x"
    assert "1" in str(ws.cell(row=2, column=1).value)   # list stringified, not crashed


# ---- resolve: per-row matched_field, literal wildcard matching --------------

def test_resolve_matched_field_reports_pid_for_id_hits():
    con = _db()
    r = resolve_from(con, "101")
    m = next(m for m in r["schedule_matches"] if m["pid"] == "101")
    assert m["matched_field"] == "pid"


def test_resolve_matched_field_reports_name_for_name_hits():
    con = _db()
    r = resolve_from(con, "Park")
    m = next(m for m in r["schedule_matches"] if m["pid"] == "101")
    assert m["matched_field"] == "agency_project_name"


def test_resolve_percent_matches_literally():
    con = _db()
    r = resolve_from(con, "%")                  # no name contains a literal '%'
    assert r["schedule_matches"] == [] and r["budget_matches"] == []


def test_escape_like_escapes_wildcards():
    assert escape_like("50% a_b\\c") == "50\\% a\\_b\\\\c"


# ---- describe_field: substring lookup -----------------------------------------

def test_describe_field_partial_name_matches():
    con = _db()
    r = describe_field_from(con, field="budget")
    assert len(r["fields"]) > 0
    assert all("budget" in (f["field_name"] + (f["display"] or "")).lower()
               for f in r["fields"])


# ---- run_sql guard: literals/comments are not scanned, wrap survives comments

def test_validate_allows_forbidden_words_inside_literals():
    validate_select("SELECT 1 WHERE 'x' LIKE '%update%'")
    validate_select("SELECT 'a;b' AS v")
    validate_select("SELECT 1 -- drop table note\n")


def test_run_sql_survives_trailing_line_comment():
    con = duckdb.connect(":memory:")
    r = run_sql_on(con, "SELECT 1 AS x -- trailing note")
    assert r["rows"] == [{"x": 1}]


def test_validate_still_blocks_real_keywords():
    with pytest.raises(ValueError):
        validate_select("SELECT 1; DROP TABLE t")
    with pytest.raises(ValueError):
        validate_select("WITH x AS (SELECT 1) UPDATE t SET a=1")


# ---- ingest: header order is validated ---------------------------------------

def test_load_raw_csv_rejects_reordered_header(tmp_path):
    con = duckdb.connect(":memory:"); schema.apply_schema(con)
    csv = tmp_path / "sched.csv"
    # the schema order with the first two columns swapped — derived from RAW_COLUMNS so
    # the test keeps detecting REORDERING (not just any stale header) as the schema evolves
    cols = list(schema.RAW_COLUMNS["raw_schedule_history"])
    cols[0], cols[1] = cols[1], cols[0]
    csv.write_text(",".join(cols) + "\n" + ",".join(["x"] * len(cols)) + "\n")
    with pytest.raises(ValueError, match="header order"):
        ingest.load_raw_csv(con, "raw_schedule_history", csv)


# ---- fms_sponsor: per-PID latest links, not global-latest ---------------------

def test_fms_sponsor_keeps_lines_with_older_links():
    con = duckdb.connect(":memory:"); _raw(con)
    # PID 555 (DOT-sponsored) funds fms F but its link last appeared at 202509,
    # one period before the global latest (202601) — it must still be attributed.
    con.execute(
        "INSERT INTO raw_project_detail (reporting_period, managing_agency, sponsor_agency,"
        " pid, fms_id, total_budget, current_phase) "
        "VALUES ('202509','DDC','DOT','555','F','400','Construction')")
    con.execute(
        "INSERT INTO raw_budget_history (managing_agency, fms_id, year_month_reported,"
        " total_budget, spend_to_date, budget_variance) "
        "VALUES ('DDC','F','202601','400','0','0')")
    materialize.materialize_all(con)
    pairs = set(con.execute("SELECT fms_id, sponsor_agency FROM fms_sponsor").fetchall())
    assert ("F", "DOT") in pairs
