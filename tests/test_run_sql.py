# tests/test_run_sql.py
import csv
import threading
from http.server import BaseHTTPRequestHandler, HTTPServer
from pathlib import Path

import duckdb
import pytest
from openpyxl import load_workbook

from od_cpd import dbio, server
from od_cpd.tools.sql import validate_select, run_sql_on


def test_validate_rejects_non_select():
    for bad in ["INSERT INTO t VALUES (1)", "DROP TABLE t", "UPDATE t SET a=1",
                "ATTACH 'x'", "PRAGMA database_list", "COPY t TO 'f'"]:
        with pytest.raises(ValueError):
            validate_select(bad)


def test_validate_rejects_multi_statement():
    with pytest.raises(ValueError):
        validate_select("SELECT 1; SELECT 2")


def test_validate_allows_select_and_with():
    validate_select("SELECT 1")
    validate_select("WITH x AS (SELECT 1) SELECT * FROM x")


def test_run_sql_inline_caps_rows():
    con = duckdb.connect(":memory:")
    con.execute("CREATE TABLE t AS SELECT * FROM range(10) AS r(n)")
    result = run_sql_on(con, "SELECT n FROM t", row_cap=3)
    assert result["truncated"] is True
    assert len(result["rows"]) == 3
    assert result["provenance"]["reproduce_sql"].startswith("SELECT")


def test_run_sql_not_truncated_when_small():
    con = duckdb.connect(":memory:")
    con.execute("CREATE TABLE t AS SELECT * FROM range(2) AS r(n)")
    result = run_sql_on(con, "SELECT n FROM t", row_cap=10)
    assert result["truncated"] is False
    assert len(result["rows"]) == 2


def _db_with_meta(latest="202601"):
    con = duckdb.connect(":memory:")
    con.execute("CREATE TABLE meta (latest_reporting_period VARCHAR)")
    con.execute("INSERT INTO meta VALUES (?)", [latest])
    return con


def test_run_sql_echoes_latest_period():
    con = _db_with_meta("202601")
    con.execute("CREATE TABLE t AS SELECT 1 AS n")
    r = run_sql_on(con, "SELECT n FROM t")
    assert r["latest_reporting_period"] == "202601"


def test_run_sql_latest_period_none_without_meta():
    con = duckdb.connect(":memory:")
    con.execute("CREATE TABLE t AS SELECT 1 AS n")
    r = run_sql_on(con, "SELECT n FROM t")
    assert r["latest_reporting_period"] is None


def test_run_sql_flags_all_history_table():
    con = _db_with_meta("202601")
    con.execute("CREATE TABLE fms_location AS SELECT 'A' AS fms_id, 'K' AS borough")
    r = run_sql_on(con, "SELECT borough, count(*) FROM fms_location GROUP BY borough")
    assert "period_basis_note" in r
    assert "fms_location" in r["period_basis_note"]
    assert "202601" in r["period_basis_note"]


def test_run_sql_no_note_for_period_scoped_query():
    con = _db_with_meta("202601")
    con.execute("CREATE TABLE raw_project_detail AS "
                "SELECT '202601' AS reporting_period, 'A' AS fms_id")
    r = run_sql_on(con, "SELECT count(*) FROM raw_project_detail "
                        "WHERE reporting_period='202601'")
    assert "period_basis_note" not in r


def test_run_sql_note_ignores_string_literal():
    con = _db_with_meta("202601")
    con.execute("CREATE TABLE t AS SELECT 'fms_location' AS s")
    r = run_sql_on(con, "SELECT s FROM t WHERE s = 'fms_location'")
    assert "period_basis_note" not in r


def test_run_sql_note_pluralizes_for_multiple_tables():
    con = _db_with_meta("202601")
    con.execute("CREATE TABLE fms_location AS SELECT 'A' AS fms_id")
    con.execute("CREATE TABLE fms_sponsor AS SELECT 'A' AS fms_id")
    r = run_sql_on(con, "SELECT * FROM fms_location JOIN fms_sponsor USING (fms_id)")
    assert "fms_location, fms_sponsor are all-history" in r["period_basis_note"]


def test_describe_rejection_steers_to_describe_table():
    import pytest
    from od_cpd.tools.sql import validate_select
    with pytest.raises(ValueError, match="describe_table"):
        validate_select("DESCRIBE schedule_history")


def test_truncated_inline_result_carries_csv_hint():
    con = duckdb.connect(":memory:")
    con.execute("CREATE TABLE meta (latest_reporting_period VARCHAR)")
    r = run_sql_on(con, "SELECT * FROM range(10)", row_cap=3)
    assert r["truncated"] is True
    assert "output='csv'" in r["truncation_note"]
    r2 = run_sql_on(con, "SELECT 1", row_cap=3)
    assert "truncation_note" not in r2


@pytest.fixture
def query_database(tmp_path, monkeypatch):
    path = tmp_path / "query.duckdb"
    with duckdb.connect(str(path)) as con:
        con.execute("CREATE TABLE projects AS SELECT 1 AS id, 'Synthetic project' AS name")
        con.execute("CREATE TABLE budgets AS SELECT 1 AS id, 100 AS amount")
    monkeypatch.setenv("OD_CPD_DB", str(path))
    monkeypatch.setenv("OD_CPD_EXPORT_DIR", str(tmp_path / "controlled exports'"))
    return path


@pytest.mark.parametrize("output", ["inline", "csv", "xlsx"])
@pytest.mark.parametrize("reader", ["read_text", "read_blob", "read_csv", "glob", "query"])
def test_run_sql_blocks_external_files_in_every_mode(query_database, tmp_path, output, reader):
    marker = tmp_path / "synthetic marker's.csv"
    marker.write_text("payload\nsynthetic_marker\n", encoding="utf-8")
    path = dbio.sql_literal(str(tmp_path / "*.csv") if reader == "glob" else str(marker))
    query = f"SELECT * FROM {reader}({path})"
    if reader == "query":
        # Dynamic SQL still encounters the engine boundary even though the outer
        # validator treats its query argument as a string literal.
        nested = dbio.sql_literal(f"SELECT * FROM read_text({path})")
        query = f"SELECT * FROM query({nested})"
    with pytest.raises(duckdb.PermissionException, match="disabled by configuration"):
        server.run_sql(query, output=output)
    assert not list((tmp_path / "controlled exports'").glob("*"))


@pytest.mark.parametrize("output", ["inline", "csv", "xlsx"])
def test_run_sql_keeps_database_joins_and_controlled_exports(query_database, output):
    query = "SELECT name, amount FROM projects JOIN budgets USING (id)"
    result = server.run_sql(query, output=output)
    assert result["provenance"]["reproduce_sql"] == query
    if output == "inline":
        assert result["rows"] == [{"name": "Synthetic project", "amount": 100}]
    elif output == "csv":
        with Path(result["file"]).open(encoding="utf-8", newline="") as stream:
            assert list(csv.reader(stream)) == [["name", "amount"], ["Synthetic project", "100"]]
    else:
        wb = load_workbook(result["file"], read_only=True)
        try:
            assert set(wb.sheetnames) == {"data", "methodology"}
            assert list(wb["data"].values) == [("name", "amount"), ("Synthetic project", 100)]
        finally:
            wb.close()


def test_run_sql_blocks_network_and_extension_autoload(query_database, tmp_path, monkeypatch):
    requests = []

    class Handler(BaseHTTPRequestHandler):
        def do_GET(self):
            requests.append(self.path)
            self.send_error(404)

        def do_HEAD(self):
            self.do_GET()

        def log_message(self, *args):
            pass

    # Both data URLs and extension repositories point only to this bounded local
    # responder. A regression can neither download code nor contact the internet.
    http = HTTPServer(("127.0.0.1", 0), Handler)
    thread = threading.Thread(target=http.serve_forever, kwargs={"poll_interval": 0.01}, daemon=True)
    thread.start()
    url = f"http://127.0.0.1:{http.server_port}"
    extension_dir = tmp_path / "isolated-extensions"
    connect = duckdb.connect

    def isolated_connect(*args, **kwargs):
        kwargs["config"] = {
            "extension_directory": str(extension_dir),
            "autoinstall_extension_repository": url,
            "custom_extension_repository": url,
            **kwargs.get("config", {}),
        }
        return connect(*args, **kwargs)

    monkeypatch.setattr(dbio.duckdb, "connect", isolated_connect)
    try:
        for output in ("inline", "csv", "xlsx"):
            with pytest.raises(duckdb.PermissionException, match="disabled by configuration"):
                server.run_sql(f"SELECT * FROM read_csv('{url}/data.csv')", output=output)
            with pytest.raises(duckdb.CatalogException, match="sqlite_scanner extension"):
                server.run_sql("SELECT * FROM sqlite_scan(':memory:', 't')", output=output)
        with dbio.ro_conn() as con:
            assert con.execute("SELECT count(*) FROM duckdb_functions() "
                               "WHERE function_name = 'sqlite_scan'").fetchone() == (0,)
        assert requests == []
        assert not list(extension_dir.rglob("*"))
    finally:
        http.shutdown()
        http.server_close()
        thread.join(timeout=2)
        assert not thread.is_alive()
