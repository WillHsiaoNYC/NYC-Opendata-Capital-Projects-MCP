# src/od_cpd/export.py
from __future__ import annotations

import csv
from datetime import date, datetime, time
from decimal import Decimal
from pathlib import Path
from uuid import uuid4

import duckdb
from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell

from .config import RUN_SQL_TIMEOUT_SECONDS, export_dir
from .dbio import interrupt_after

# Cell types openpyxl writes natively; anything else (DuckDB LIST/STRUCT → Python
# list/dict) raises ValueError("Cannot convert ... to Excel") and must be stringified.
_XLSX_NATIVE = (str, int, float, bool, Decimal, datetime, date, time, type(None))


def _unique_name() -> str:
    """A fresh filename per export — a fixed name silently overwrites the previous
    export, invalidating any path a client is still holding."""
    return f"export_{datetime.now():%Y%m%d_%H%M%S}_{uuid4().hex[:6]}"


def _ensure_dir(path: Path) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    return path


def _xlsx_value(sheet, value):
    """Keep SQL strings literal while retaining Excel's numeric and date types."""
    if not isinstance(value, _XLSX_NATIVE):
        value = str(value)
    if isinstance(value, str):
        cell = WriteOnlyCell(sheet, value=value)
        # openpyxl otherwise interprets leading '=' and strings such as '#N/A'.
        cell.data_type = "s"
        return cell
    return value


def write_csv(con: duckdb.DuckDBPyConnection, select_sql: str, out: Path | None = None,
              timeout: int = RUN_SQL_TIMEOUT_SECONDS) -> Path:
    """Export faithful CSV values, including formula-like text without alteration.

    CSV has no cell types; use XLSX when strings must stay literal in a spreadsheet.
    """
    out = _ensure_dir(out or (export_dir() / f"{_unique_name()}.csv"))
    # timer guards the export like the inline path — an uncapped runaway query (e.g. an
    # accidental self cross-join) would otherwise hang the stdio server indefinitely.
    timer = interrupt_after(con, timeout)
    try:
        # Let DuckDB render values before Python conversion can lose information
        # (e.g. interval months, timestamp precision, or nested type notation).
        # The newline keeps a trailing SQL comment from swallowing the wrapper.
        cur = con.execute(f"SELECT COLUMNS(*)::VARCHAR FROM ({select_sql}\n) AS _csv")
        # Query connections cannot access files. Serialize in application code;
        # quoting non-NULL values preserves the distinction between NULL and ''.
        with out.open("w", encoding="utf-8", newline="") as stream:
            csv.writer(stream, lineterminator="\n").writerow([d[0] for d in cur.description])
            writer = csv.writer(stream, lineterminator="\n", quoting=csv.QUOTE_NOTNULL)
            while batch := cur.fetchmany(5000):
                writer.writerows(batch)
    finally:
        timer.cancel()
    return out


def write_xlsx(con: duckdb.DuckDBPyConnection, select_sql: str, provenance: dict,
               out: Path | None = None, timeout: int = RUN_SQL_TIMEOUT_SECONDS) -> Path:
    out = _ensure_dir(out or (export_dir() / f"{_unique_name()}.xlsx"))
    # write_only streams rows instead of building one Cell object per cell, and fetchmany
    # caps the Python-side buffer — the export path has no row cap, so a full-table
    # export would otherwise spike the long-running server process by hundreds of MB.
    wb = Workbook(write_only=True)
    ws = wb.create_sheet("data")
    # timer guards query + streaming fetch like the inline path — an uncapped runaway
    # query would otherwise hang the stdio server indefinitely. (No paren wrap here, so
    # a trailing `--` comment is already harmless.)
    timer = interrupt_after(con, timeout)
    try:
        cur = con.execute(select_sql)
        ws.append([_xlsx_value(ws, d[0]) for d in cur.description])
        while batch := cur.fetchmany(5000):
            for r in batch:
                ws.append([_xlsx_value(ws, v) for v in r])
    finally:
        timer.cancel()
    meth = wb.create_sheet("methodology")
    for k, v in provenance.items():
        meth.append([_xlsx_value(meth, k), _xlsx_value(meth, str(v))])
    wb.save(out)
    return out
